"""
Gerador WP Excel — padrão Cast Group
Produz o Work Package com fases SAP Activate, recursos e totalizadores
"""
import io
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as num_styles)
from openpyxl.utils import get_column_letter

AZUL_FILL    = PatternFill("solid", fgColor="1F4E79")
AZUL_CL_FILL = PatternFill("solid", fgColor="2E75B6")
CINZA_FILL   = PatternFill("solid", fgColor="D9E1F2")
VERDE_FILL   = PatternFill("solid", fgColor="E2EFDA")
AMARELO_FILL = PatternFill("solid", fgColor="FFF2CC")
LARANJA_FILL = PatternFill("solid", fgColor="FCE4D6")
BRANCO_FILL  = PatternFill("solid", fgColor="FFFFFF")

FONT_TITLE  = Font(name="Calibri", size=16, bold=True, color="1F4E79")
FONT_HEAD   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_FASE   = Font(name="Calibri", size=10, bold=True, color="1F4E79")
FONT_BODY   = Font(name="Calibri", size=10)
FONT_TOTAL  = Font(name="Calibri", size=10, bold=True)
FONT_SMALL  = Font(name="Calibri", size=9, color="595959")

BORDER_THIN = Border(
    left  =Side(style='thin',  color="BFBFBF"),
    right =Side(style='thin',  color="BFBFBF"),
    top   =Side(style='thin',  color="BFBFBF"),
    bottom=Side(style='thin',  color="BFBFBF"),
)
BORDER_MED = Border(
    left  =Side(style='medium',color="595959"),
    right =Side(style='medium',color="595959"),
    top   =Side(style='medium',color="595959"),
    bottom=Side(style='medium',color="595959"),
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

def cell_style(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    return c

def generate_wp(resources: list, payload: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "WP_RFP"
    ws.sheet_view.showGridLines = False

    # ── Título ──
    ws.merge_cells("A1:H1")
    ws["A1"].value     = "WORK PACKAGE — SIL-PROPOSTA"
    ws["A1"].font      = FONT_TITLE
    ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 30

    # ── Constantes ──
    HORAS_DIA = 8
    semanas_max = 6

    # ── Cabeçalho semanas ──
    header_row = 3
    ws.cell(row=header_row, column=1, value="Frente").font     = FONT_HEAD
    ws.cell(row=header_row, column=1).fill                     = AZUL_CL_FILL
    ws.cell(row=header_row, column=1).alignment                = CENTER
    ws.cell(row=header_row, column=2, value="Recurso").font    = FONT_HEAD
    ws.cell(row=header_row, column=2).fill                     = AZUL_CL_FILL
    ws.cell(row=header_row, column=2).alignment                = CENTER
    ws.cell(row=header_row, column=3, value="Nível").font      = FONT_HEAD
    ws.cell(row=header_row, column=3).fill                     = AZUL_CL_FILL
    ws.cell(row=header_row, column=3).alignment                = CENTER

    for s in range(1, semanas_max + 1):
        col = 3 + s
        c   = ws.cell(row=header_row, column=col, value=f"Sem {s}")
        c.font      = FONT_HEAD
        c.fill      = AZUL_CL_FILL
        c.alignment = CENTER

    col_td = 3 + semanas_max + 1
    col_th = col_td + 1

    ws.cell(row=header_row, column=col_td, value="Total Dias").font      = FONT_HEAD
    ws.cell(row=header_row, column=col_td).fill                          = AZUL_FILL
    ws.cell(row=header_row, column=col_td).alignment                     = CENTER
    ws.cell(row=header_row, column=col_th, value="Total Horas").font     = FONT_HEAD
    ws.cell(row=header_row, column=col_th).fill                          = AZUL_FILL
    ws.cell(row=header_row, column=col_th).alignment                     = CENTER

    ws.row_dimensions[header_row].height = 24

    # ── Fases SAP Activate ──
    fases = [
        ("Prepare",             "Aprovação e planejamento do projeto",      CINZA_FILL),
        ("Explore",             "Levantamento, especificação funcional",     VERDE_FILL),
        ("Realize",             "Desenvolvimento, configuração e testes",    AMARELO_FILL),
        ("Deploy",              "KT AMS, cutover e go-live",                LARANJA_FILL),
        ("Suporte pós Go-Live", "Suporte e acompanhamento inicial",         CINZA_FILL),
    ]

    current_row = header_row + 1

    # ── Distribuição por fase ──
    # Mapeia recursos para fases baseado no tipo de frente
    def get_distribution(frente: str, total_dias: int) -> dict:
        """Distribui os dias do recurso pelas semanas"""
        dist = {}
        if frente in ("SD", "FI"):
            dist[1] = min(5, total_dias)
            rem = total_dias - dist.get(1,0)
            if rem > 0: dist[2] = min(5, rem)
            rem -= dist.get(2,0)
            if rem > 0: dist[3] = min(5, rem)
        elif "ABAP" in frente:
            dist[2] = min(5, total_dias)
            rem = total_dias - dist.get(2,0)
            if rem > 0: dist[3] = min(5, rem)
            rem -= dist.get(3,0)
            if rem > 0: dist[4] = min(rem, 5)
        elif frente == "GP":
            for s in range(1, min(total_dias+1, semanas_max+1)):
                dist[s] = 1
        return dist

    # Se não há recursos, usar padrão
    if not resources:
        resources = [
            {"frente":"SD",     "nivel":"Senior","dias":9},
            {"frente":"FI",     "nivel":"Senior","dias":11},
            {"frente":"GP",     "nivel":"Senior","dias":4},
            {"frente":"ABAP 1", "nivel":"Senior","dias":14},
            {"frente":"ABAP 2", "nivel":"Senior","dias":14},
            {"frente":"ABAP 3", "nivel":"Senior","dias":14},
        ]

    # ── Fase Realize (principal) — listar recursos ──
    fase_header_row = current_row
    ws.merge_cells(start_row=fase_header_row, start_column=1,
                   end_row=fase_header_row,   end_column=col_th)
    c = ws.cell(row=fase_header_row, column=1, value="⬛  Recursos do Projeto")
    c.font      = FONT_FASE
    c.fill      = CINZA_FILL
    c.alignment = LEFT
    ws.row_dimensions[fase_header_row].height = 20
    current_row += 1

    total_horas_proj = 0

    for rec in resources:
        frente = rec.get("frente","")
        nivel  = rec.get("nivel","Senior")
        dias   = rec.get("dias", 10)
        dist   = get_distribution(frente, dias)

        row = current_row
        ws.cell(row=row, column=1, value=frente).font      = FONT_BODY
        ws.cell(row=row, column=1).alignment               = LEFT
        ws.cell(row=row, column=1).border                  = BORDER_THIN
        ws.cell(row=row, column=2, value=frente).font      = FONT_BODY
        ws.cell(row=row, column=2).alignment               = LEFT
        ws.cell(row=row, column=2).border                  = BORDER_THIN
        ws.cell(row=row, column=3, value=nivel).font       = FONT_SMALL
        ws.cell(row=row, column=3).alignment               = CENTER
        ws.cell(row=row, column=3).border                  = BORDER_THIN

        total_dias_rec = 0
        for s in range(1, semanas_max + 1):
            col = 3 + s
            val = dist.get(s, 0) or ""
            c = ws.cell(row=row, column=col, value=val if val else None)
            c.alignment = CENTER
            c.border    = BORDER_THIN
            if val:
                c.fill = VERDE_FILL if frente not in ("GP",) else CINZA_FILL
                total_dias_rec += val

        horas_rec = total_dias_rec * HORAS_DIA
        total_horas_proj += horas_rec

        c_td = ws.cell(row=row, column=col_td, value=total_dias_rec)
        c_td.font      = FONT_TOTAL
        c_td.alignment = CENTER
        c_td.border    = BORDER_THIN
        c_td.fill      = CINZA_FILL

        c_th = ws.cell(row=row, column=col_th, value=horas_rec)
        c_th.font      = FONT_TOTAL
        c_th.alignment = CENTER
        c_th.border    = BORDER_THIN
        c_th.fill      = CINZA_FILL

        ws.row_dimensions[row].height = 20
        current_row += 1

    # ── KT AMS (obrigatório) ──
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row,   end_column=col_th)
    c = ws.cell(row=current_row, column=1, value="⬛  Deploy — KT AMS (obrigatório)")
    c.font      = FONT_FASE
    c.fill      = LARANJA_FILL
    c.alignment = LEFT
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    ws.cell(row=current_row, column=1, value="KT AMS").font      = FONT_BODY
    ws.cell(row=current_row, column=1).border                    = BORDER_THIN
    ws.cell(row=current_row, column=2, value="Transferência de conhecimento").font = FONT_SMALL
    ws.cell(row=current_row, column=2).border                    = BORDER_THIN
    ws.cell(row=current_row, column=3, value="—").alignment      = CENTER
    ws.cell(row=current_row, column=3).border                    = BORDER_THIN
    for s in range(1, semanas_max + 1):
        ws.cell(row=current_row, column=3+s).border = BORDER_THIN
    ws.cell(row=current_row, column=3+semanas_max-1, value=2).fill = LARANJA_FILL
    ws.cell(row=current_row, column=3+semanas_max-1).border      = BORDER_THIN
    ws.cell(row=current_row, column=col_td, value=2).font        = FONT_TOTAL
    ws.cell(row=current_row, column=col_td).alignment            = CENTER
    ws.cell(row=current_row, column=col_td).border               = BORDER_THIN
    ws.cell(row=current_row, column=col_th, value=16).font       = FONT_TOTAL
    ws.cell(row=current_row, column=col_th).alignment            = CENTER
    ws.cell(row=current_row, column=col_th).border               = BORDER_THIN
    total_horas_proj += 16
    current_row += 2

    # ── TOTAL GERAL ──
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row,   end_column=col_td-1)
    c = ws.cell(row=current_row, column=1, value="TOTAL GERAL DO PROJETO")
    c.font      = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill      = AZUL_FILL
    c.alignment = LEFT
    c.border    = BORDER_MED

    c_th = ws.cell(row=current_row, column=col_th, value=total_horas_proj)
    c_th.font      = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    c_th.fill      = AZUL_FILL
    c_th.alignment = CENTER
    c_th.border    = BORDER_MED
    ws.row_dimensions[current_row].height = 28

    current_row += 2

    # ── CUSTO INTERNO (horas pré-venda) ──
    presale = payload.get("presale_hours", 0) or payload.get("hours_presale", 0) or 0
    if presale and presale > 0:
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row,   end_column=col_th)
        c = ws.cell(row=current_row, column=1,
            value=f"⚙  CUSTO INTERNO — Horas de pré-venda: {presale}h "
                  f"(não incluídas na proposta ao cliente)")
        c.font      = Font(name="Calibri", size=9, italic=True, color="7F7F7F")
        c.alignment = LEFT
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row,   end_column=col_th)
        pct = round((presale / total_horas_proj) * 100, 1) if total_horas_proj else 0
        c = ws.cell(row=current_row, column=1,
            value=f"   Total real (faturável + pré-venda): {total_horas_proj + presale}h  |  "
                  f"Custo oculto: {pct}%")
        c.font      = Font(name="Calibri", size=9, italic=True, color="7F7F7F")
        c.alignment = LEFT

    # ── Largura das colunas ──
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    for s in range(1, semanas_max + 1):
        ws.column_dimensions[get_column_letter(3+s)].width = 8
    ws.column_dimensions[get_column_letter(col_td)].width = 12
    ws.column_dimensions[get_column_letter(col_th)].width = 13

    ws.freeze_panes = "D4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
